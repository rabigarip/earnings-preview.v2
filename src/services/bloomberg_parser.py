"""
Bloomberg MODL Excel template parser.

Parses the "Single Period" or "Values" sheet from Bloomberg Terminal's
MODL export to extract broker-by-broker consensus estimates.

Bloomberg is PRIMARY when uploaded — MS becomes fallback.
"""
from __future__ import annotations
import io
import logging
from datetime import datetime, timezone

log = logging.getLogger(__name__)

# Bloomberg field expression → our metric key mapping
_METRIC_MAP = {
    "IS_COMP_SALES": "revenue",
    "IS_COMPARABLE_EBITDA": "ebitda",
    "IS_COMPARABLE_EBIT": "ebit",
    "IS_COMP_NET_INCOME_ADJUST_OLD": "net_income",
    "IS_COMP_EPS_ADJUSTED_OLD": "eps",
    "EBITDA_TO_REVENUE": "ebitda_margin",
}

# Friendly names for display
_METRIC_LABELS = {
    "revenue": "Revenue",
    "ebitda": "EBITDA",
    "ebit": "EBIT",
    "net_income": "Net Income",
    "eps": "EPS",
    "ebitda_margin": "EBITDA Margin (%)",
}


def parse_bloomberg_excel(file_bytes: bytes, filename: str = "") -> dict:
    """Parse Bloomberg MODL template Excel.

    Args:
        file_bytes: raw Excel file content
        filename: original filename for logging

    Returns:
        {
            'ticker': str,          # e.g. "EEC AB"
            'period': str,          # e.g. "2026:Q1"
            'currency': str,        # e.g. "SAR"
            'consensus': {metric: {mean, low, high, median}},
            'brokers': [{name, analyst, estimates: {metric: value}}],
            'segments': {segment_name: {metric: value}},
            'metadata': {template_type, extracted_at, filename},
            'warnings': [str],
        }
    """
    import openpyxl

    result = {
        "ticker": "",
        "period": "",
        "currency": "",
        "consensus": {},
        "brokers": [],
        "segments": {},
        "metadata": {
            "template_type": "MODL",
            "extracted_at": datetime.now(timezone.utc).isoformat(),
            "filename": filename,
        },
        "warnings": [],
    }

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        result["warnings"].append(f"Cannot open Excel: {e}")
        return result

    # Find the right sheet
    sheet_name = None
    for candidate in ["Single Period", "Values", "Sheet1"]:
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if not sheet_name and wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    if not sheet_name:
        result["warnings"].append("No valid sheet found")
        return result

    ws = wb[sheet_name]

    # Row 1: Company name
    # Row 2: "EEC AB Equity    Period:2026:Q1    Currency:SAR"
    # Row 3: Headers (Metric, [empty cols], Mean, Low, High, Median, Broker1, Broker2, ...)
    # Row 4: Analyst names (Analyst, [empty cols], [empty], [empty], [empty], [empty], Name1, Name2, ...)

    row2_val = str(ws.cell(row=2, column=1).value or "")
    # Parse ticker, period, currency from row 2
    parts = row2_val.split()
    if len(parts) >= 2:
        result["ticker"] = f"{parts[0]} {parts[1]}".replace("Equity", "").strip()
    for part in parts:
        if "Period:" in part or (len(part) == 7 and ":" in part):
            result["period"] = part.replace("Period:", "")
        if "Currency:" in part:
            result["currency"] = part.replace("Currency:", "")
    # Also try splitting by spaces
    for i, part in enumerate(parts):
        if part == "Period:" and i + 1 < len(parts):
            result["period"] = parts[i + 1]
        if part == "Currency:" and i + 1 < len(parts):
            result["currency"] = parts[i + 1]

    # Parse row 2 more carefully
    import re
    period_match = re.search(r"Period:(\S+)", row2_val)
    if period_match:
        result["period"] = period_match.group(1)
    currency_match = re.search(r"Currency:(\S+)", row2_val)
    if currency_match:
        result["currency"] = currency_match.group(1)

    # Row 3: Find column positions for Mean, Low, High, Median, and broker names
    header_row = 3
    broker_start_col = None
    consensus_cols = {}  # {"mean": col, "low": col, ...}
    broker_names = []

    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=col).value or "").strip()
        val_lower = val.lower()
        if "mean" in val_lower:
            consensus_cols["mean"] = col
        elif "low" in val_lower and "consensus" in val_lower:
            consensus_cols["low"] = col
        elif "high" in val_lower and "consensus" in val_lower:
            consensus_cols["high"] = col
        elif "median" in val_lower:
            consensus_cols["median"] = col
        elif val and val not in ("Metric", "") and col > 4 and col not in consensus_cols.values():
            if broker_start_col is None:
                broker_start_col = col
            broker_names.append((col, val))

    # Row 4: Analyst names for each broker column
    analyst_names = {}
    for col, broker_name in broker_names:
        analyst = ws.cell(row=4, column=col).value
        if analyst and str(analyst).strip():
            analyst_names[col] = str(analyst).strip()

    # Parse data rows (row 5 onward)
    for row_idx in range(5, ws.max_row + 1):
        metric_name = str(ws.cell(row=row_idx, column=1).value or "").strip()
        field_expr = str(ws.cell(row=row_idx, column=2).value or "").strip()

        if not metric_name or not field_expr:
            continue

        # Map to our metric key
        metric_key = _METRIC_MAP.get(field_expr)
        if not metric_key:
            continue

        # Extract consensus values
        consensus = {}
        for label, col in consensus_cols.items():
            val = ws.cell(row=row_idx, column=col).value
            if val is not None and val != "":
                try:
                    consensus[label] = float(val)
                except (TypeError, ValueError):
                    pass

        if consensus:
            result["consensus"][metric_key] = {
                "label": _METRIC_LABELS.get(metric_key, metric_name),
                **consensus,
            }

        # Extract individual broker estimates
        for col, broker_name in broker_names:
            val = ws.cell(row=row_idx, column=col).value
            if val is not None and val != "":
                try:
                    float_val = float(val)
                except (TypeError, ValueError):
                    continue

                # Find or create broker entry
                broker_entry = None
                for b in result["brokers"]:
                    if b["name"] == broker_name:
                        broker_entry = b
                        break
                if not broker_entry:
                    broker_entry = {
                        "name": broker_name,
                        "analyst": analyst_names.get(col, ""),
                        "estimates": {},
                    }
                    result["brokers"].append(broker_entry)
                broker_entry["estimates"][metric_key] = float_val

    if not result["consensus"]:
        result["warnings"].append("No consensus data found in template")

    log.info("Bloomberg parsed: %s %s %s — %d metrics, %d brokers",
             result["ticker"], result["period"], result["currency"],
             len(result["consensus"]), len(result["brokers"]))

    return result
